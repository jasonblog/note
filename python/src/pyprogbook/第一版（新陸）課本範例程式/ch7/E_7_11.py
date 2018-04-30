#E_7_11.py功能: 使用openpyxl讀寫excel檔案
from openpyxl import load_workbook
import statistics as st
wb = load_workbook('./file/oil_statistics.xlsx') 
ws1 = wb.active 
rows=ws1.get_highest_row()
cols=ws1.get_highest_column()
print('資料總列數 = ',rows)
print('資料總行數 = ',cols)
ws2 = wb.create_sheet()
ws2.title = '原油價格'
ws2.sheet_properties.tabColor='F4A460'
ws2['A1'] = '統計量'
ws2['A2'] = '平均值'
ws2['A3'] = '標準差'
ws2['A4'] = '最高價'
ws2['A5'] = '最低價'
ws2['B1'] = '西德州'
ws2['C1'] = '杜拜'
ws2['D1'] = '布蘭特'
li1=[]
for i in range(2,cols+1):
    for j in range(2, rows+1):
        li1.append(ws1.cell(row=j, column=i).value)
    ws2.cell(row=2,column=i).value = round(st.mean(li1),2)
    ws2.cell(row=3,column=i).value = round(st.stdev(li1),2)
    ws2.cell(row=4,column=i).value = max(li1)
    ws2.cell(row=5,column=i).value = min(li1)
    li1=[]
wb.save('./file/oil_statistics.xlsx')